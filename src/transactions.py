"""Utility functions for working with banking transactions."""

from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
import csv
import json
import re
from typing import Any, Iterable, Iterator

from .widget import get_date, mask_account_card


try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None  # type: ignore

Transaction = dict[str, Any]


def load_transactions(file_path: str | Path) -> list[Transaction]:
    """Return a list of transactions loaded from *file_path*.

    The loader supports JSON, CSV and (optionally) XLSX files. Regardless of the
    source file format, the returned dictionaries share the same structure:

    ``{"operationAmount": {"amount": str, "currency": {"code": str, "name": str}}}``.
    """

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл {path} не найден")

    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_from_json(path)
    if suffix == ".csv":
        return _load_from_csv(path)
    if suffix in {".xlsx", ".xls"}:
        return _load_from_xlsx(path)
    raise ValueError(f"Неизвестный формат файла: {path.suffix}")


def _load_from_json(path: Path) -> list[Transaction]:
    with path.open(encoding="utf-8") as fp:
        data = json.load(fp)
    if not isinstance(data, list):
        raise ValueError("Ожидался список транзакций в JSON-файле")
    return [dict(item) for item in data]


def _load_from_csv(path: Path) -> list[Transaction]:
    with path.open(encoding="utf-8") as fp:
        reader = csv.DictReader(fp)
        transactions: list[Transaction] = []
        for row in reader:
            transactions.append(
                {
                    "id": _safe_int(row.get("id")),
                    "date": row.get("date", ""),
                    "description": row.get("description", ""),
                    "from": row.get("from") or None,
                    "to": row.get("to") or None,
                    "state": row.get("state", ""),
                    "operationAmount": {
                        "amount": row.get("amount", ""),
                        "currency": {
                            "code": row.get("currency_code", ""),
                            "name": row.get("currency_name", ""),
                        },
                    },
                }
            )
    return transactions


def _load_from_xlsx(path: Path) -> list[Transaction]:
    if load_workbook is None:  # pragma: no cover - depends on optional package
        raise ImportError(
            "Для чтения XLSX-файлов установите зависимость 'openpyxl'."
        )

    workbook = load_workbook(path)
    sheet = workbook.active
    header: list[str] = [str(cell.value) for cell in next(_iter_rows(sheet.iter_rows(min_row=1, max_row=1)))]
    expected = {
        "id",
        "date",
        "description",
        "from",
        "to",
        "state",
        "amount",
        "currency_code",
        "currency_name",
    }
    if set(header) != expected:
        raise ValueError("Неожиданная структура XLSX-файла")

    transactions: list[Transaction] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(header, (value or "" for value in row)))
        transactions.append(
            {
                "id": _safe_int(row_data.get("id")),
                "date": str(row_data.get("date", "")),
                "description": str(row_data.get("description", "")),
                "from": str(row_data.get("from")) if row_data.get("from") else None,
                "to": str(row_data.get("to")) if row_data.get("to") else None,
                "state": str(row_data.get("state", "")),
                "operationAmount": {
                    "amount": str(row_data.get("amount", "")),
                    "currency": {
                        "code": str(row_data.get("currency_code", "")),
                        "name": str(row_data.get("currency_name", "")),
                    },
                },
            }
        )
    return transactions


def filter_transactions_by_description(transactions: Iterable[Transaction], search: str) -> list[Transaction]:
    """Return transactions that contain *search* in their description."""

    if not search:
        return list(transactions)

    pattern = re.compile(re.escape(search), flags=re.IGNORECASE)
    return [transaction for transaction in transactions if pattern.search(transaction.get("description", ""))]


def count_transactions_by_categories(
    transactions: Iterable[Transaction], categories: Iterable[str]
) -> dict[str, int]:
    """Return statistics for *categories* based on the description field."""

    normalized_categories: dict[str, str] = {}
    for category in categories:
        trimmed = category.strip()
        if trimmed:
            normalized_categories.setdefault(trimmed.lower(), trimmed)

    counter: Counter[str] = Counter()
    for transaction in transactions:
        description = transaction.get("description", "").lower()
        for normalized in normalized_categories:
            if normalized in description:
                counter[normalized] += 1

    return {original: counter.get(normalized, 0) for normalized, original in normalized_categories.items()}


def filter_transactions_by_status(transactions: Iterable[Transaction], status: str) -> list[Transaction]:
    """Return transactions that match the provided *status* (case-insensitive)."""

    normalized = status.strip().lower()
    return [transaction for transaction in transactions if transaction.get("state", "").lower() == normalized]


def sort_transactions_by_date(transactions: Iterable[Transaction], ascending: bool = False) -> list[Transaction]:
    """Sort transactions by the ``date`` field and return a new list."""

    def parse_date(transaction: Transaction) -> datetime:
        value = str(transaction.get("date", ""))
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return datetime.min

    return sorted(transactions, key=parse_date, reverse=not ascending)


def filter_transactions_by_currency(transactions: Iterable[Transaction], currency_code: str) -> list[Transaction]:
    """Filter transactions by the currency code."""

    normalized = currency_code.strip().upper()
    result: list[Transaction] = []
    for transaction in transactions:
        currency = transaction.get("operationAmount", {}).get("currency", {})
        if str(currency.get("code", "")).upper() == normalized:
            result.append(transaction)
    return result


def _safe_int(value: Any) -> int | None:
    """Safely cast a value to int when possible."""

    try:
        return int(value) if value not in {None, ""} else None
    except (TypeError, ValueError):
        return None


def _iter_rows(iterator: Iterator[Any]) -> Iterator[Any]:
    """Helper that returns the iterator unchanged (workaround for coverage)."""

    return iterator

def format_transaction(transaction: Transaction) -> str:
    """Format transaction details for console output."""

    date_value = transaction.get("date", "")
    formatted_date = get_date(str(date_value)) if date_value else ""
    description = transaction.get("description", "")
    lines = [f"{formatted_date} {description}".strip()]

    sender = transaction.get("from")
    recipient = transaction.get("to")
    if sender and recipient:
        lines.append(f"{_format_party(sender)} -> {_format_party(recipient)}")
    elif sender:
        lines.append(_format_party(sender))
    elif recipient:
        lines.append(_format_party(recipient))

    amount_info = transaction.get("operationAmount", {})
    currency_info = amount_info.get("currency", {})
    amount = amount_info.get("amount", "")
    currency_name = currency_info.get("name", "")
    lines.append(f"Сумма: {amount} {currency_name}".strip())

    return "\n".join(lines) + "\n"


def _format_party(value: str) -> str:
    """Return masked representation for account/card if digits are present."""

    if any(char.isdigit() for char in value):
        return mask_account_card(value)
    return value
